import pandas as pd
import openpyxl
from openpyxl import load_workbook
data1 = {'IP': '49.86.9.86', 'Port': '9999', '地址': ['中国  江苏  扬州'], '运营商': ['电信']}
data2 = {'IP': '49.86.9.12', 'Port': '9982', '地址': ['中国  江苏  扬州'], '运营商': ['电信']}
data3 = ['49.86.9.86','9999','中国  江苏  扬州','电信']
data4 = ['49.86.9.86','99','中国  江苏  扬州','电信']
def write_excel_xlsx(path, sheet_name, value):
    index = len(value)
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = sheet_name
    for i in range(0, index):
        for j in range(0, len(value[i])):
            sheet.cell(row=i+1, column=j+1, value=str(value[i][j]))
    workbook.save(path)

path = 'D:/test/Code/exercise/reptile/data/reptile_verify.xlsx'
